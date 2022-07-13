from turtle import title
from unicodedata import name
from flask import Flask, flash, render_template, request, redirect, url_for, session
from sheets import update_details
from scraper import get_amazon_price_by_link, get_price_name, get_walamart_price, get_amazon_price, update_db
from squares import add_to_square
from marketplace import uploadToMP, publish
from printzpls import printlabel, open_acrobat_print
from openpyxl import load_workbook
from ig import postInstagram
import os
import pandas as pd
from scraper_test import get_products_upc, get_products_category, get_products_tcin, get_tcin_upc, insert_into_table

from flask_session import Session
import datetime
import json
import asyncio
import urllib.request
import sqlite3
import threading
import pyppeteer
import bs4
import requests
from notifypy import Notify

import csv
import time

app = Flask(__name__)
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

timestr = time.strftime("%Y%m%d")

SITE_ROOT = os.path.realpath(os.path.dirname(__file__))
json_url = os.path.join(SITE_ROOT, "static/data", "categories.json")
config_url = os.path.join(SITE_ROOT, "static/data", "config.json")
data_date_url = os.path.join(SITE_ROOT, "static/data", "data" + timestr + ".json")
data_url = os.path.join(SITE_ROOT, "static/data", "data.json")
data_ig_url = os.path.join(SITE_ROOT, "static/data", "data_ig.json")


upcDetails={}
isAvailable = 1


table_name = ""

today = datetime.datetime.today()
print(today)

API_URL1 = "https://redsky.target.com/redsky_aggregations/v1/web/plp_search_v1"
API_URL2 = "https://redsky.target.com/redsky_aggregations/v1/web/pdp_client_v1"
API_KEY = "9f36aeafbe60771e321a7cc95a78140772ab3e96"

def notify(title, msg):
    notification = Notify()
    notification.title = title
    notification.message = msg
    notification.send()

def testing():
    today = datetime.datetime.today()
    print(today)
    # Testing weekly
    if today.weekday() == 0:
        print("updating database....")
        notify('Updating DB....', "Today is Monday, so system is doing updating database")
        conn = sqlite3.connect('mydb.db')
        cur = conn.cursor()
        # start - updating database
        sql_query = "SELECT * FROM history WHERE id=(SELECT MAX(id) FROM history)"
        results = cur.execute(sql_query).fetchall()
        if len(results):
            last_scrapped_date = results[0][1]
        else:
            last_scrapped_date = '0000-00-00'
        diff_d = (int(today.strftime("%d")) - int(last_scrapped_date.split('-')[2]) + 30) % 30
        if diff_d >= 7:
            categories = json.load(open(os.path.join("categories.json")))
            get_products_category(categories)
            sql_query = "INSERT INTO history (date) VALUES (" + '"' + today.strftime("%Y-%m-%d") + '"' + ")"
            cur.execute(sql_query)
            conn.commit()
        # end - updating database
        # start - testing product is available
        sql_query = "SELECT * FROM products_manual ORDER BY update_date DESC"
        results = cur.execute(sql_query).fetchall()
        for row in results:
            if row[13] != today.strftime('%Y-%m-%d'):
                sql_query = "UPDATE products_manual SET is_available=0 WHERE id=" + str(row[0])
                cur.execute(sql_query)
                conn.commit()
        today_m = int(today.strftime("%m"))
        sql_query = "SELECT tcin, open_date, update_date, last_sold, is_available FROM products_manual"
        results = cur.execute(sql_query).fetchall()
        for row in results:
            if (today_m + 12 - int(row[2].split('-')[1])) % 12 > 6 and not row[0][4]:
                # deleting unavailable product from database
                sql_query = "DELETE FROM products_manual WHERE tcin=" + '"' + row[0] + '"'
                cur.execute(sql_query)
                conn.commit()
            elif row[3] != "0000-00-00" and (today_m + 12 - int(row[3].split('-')[1])) % 12 > 3:
                # not sold in 3 months product
                sql_query = "DELETE FROM products_manual WHERE tcin=" + '"' + row[0] + '"'
                cur.execute(sql_query)
                conn.commit()
        for category in categories:
            products_table = category['name'] + "_products"
            products_table = products_table.replace(" ", "")
            sql_query = "SELECT * FROM " + products_table + " ORDER BY update_date DESC"
            results = cur.execute(sql_query).fetchall()
            for row in results:
                if row[13] != today.strftime('%Y-%m-%d'):
                    sql_query = "UPDATE " + products_table + " SET is_available=0 WHERE id=" + str(row[0])
                    cur.execute(sql_query)
                    conn.commit()
            today_m = int(today.strftime("%m"))
            sql_query = "SELECT tcin, open_date, update_date, last_sold, is_available FROM " + products_table
            results = cur.execute(sql_query).fetchall()
            for row in results:
                if (today_m + 12 - int(row[2].split('-')[1])) % 12 > 6 and not row[0][4]:
                    # deleting unavailable product from database
                    sql_query = "DELETE FROM " + products_table + " WHERE tcin=" + '"' + row[0] + '"'
                    cur.execute(sql_query)
                    conn.commit()
                elif row[3] != "0000-00-00" and (today_m + 12 - int(row[3].split('-')[1])) % 12 > 3:
                    # not sold in 3 months product
                    sql_query = "DELETE FROM " + products_table + " WHERE tcin=" + '"' + row[0] + '"'
                    cur.execute(sql_query)
                    conn.commit()
                    

        # end - testing product is available
        conn.close()

    # end - testing


@app.route('/add', methods = ['GET', 'POST'])
def add_produtcs():
    conn = sqlite3.connect('mydb.db')
    cur = conn.cursor()
    sql_query = "CREATE TABLE IF NOT EXISTS products_manual" + \
                    " ( id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, " + \
                    "url TEXT," + \
                    "tcin TEXT," + \
                    "upc TEXT," + \
                    "name TEXT," + \
                    "description TEXT," + \
                    "image TEXT," + \
                    "category TEXT," + \
                    "price TEXT," + \
                    "disc TEXT," + \
                    "stock TEXT," + \
                    "employee TEXT," + \
                    "open_date TEXT," + \
                    "update_date TEXT," + \
                    "is_available integer DEFAULT 1," + \
                    "last_sold TEXT," + \
                    "last_price TEXT);"
    global upcDetails
    try:
        vender = upcDetails["employee"]
    except:
        vender = ""
    
    if request.method == "POST":
        # try:
        #     stock=request.form["Stock"]
        # except:
        #     stock=0
        
        try:
            printlabels = int(request.form["printlabels"])
        except:
            printlabels = 1
            
        employee=request.form["VendorName"]  
        
        if request.form["btn"] == 'Print New Label':
            product_name = request.form["Name"]  
            product_price = request.form["Price"]  
            # description = request.form["Description"] 
            #category = request.form["Category"] 
            # category = "Miscellaneous"
            # condition = request.form["Condition"]             
            # upc = datetime.datetime.now().strftime('%Y%m%d%H%M')
            upc = request.form["upc"]
            # stock = request.form["Stock"] 
            try:
                imgpath = request.files["filename"].filename
            except:
                imgpath = ""
            
            sql_query = "INSERT INTO " + table_name + "_scaned" + \
                    " (upc, name, image, price, employee) VALUES (" + \
                    '"' + upc + '", ' + \
                    '"' + product_name + '", ' + \
                    '"' + imgpath + '", ' + \
                    str(product_price) + ", " + \
                    '"' + employee + '");'
            print(sql_query)
            cur.execute(sql_query)
            conn.commit()
            sql_query = "INSERT INTO " + table_name + "_printed" + \
                    " (upc, name, image, price, employee ) VALUES (" + \
                    '"' + upc + '", ' + \
                    '"' + product_name + '", ' + \
                    '"' + imgpath + '", ' + \
                    str(product_price) + ", " + \
                    '"' + employee + '");'
            print(sql_query)
            cur.execute(sql_query)
            conn.commit()
            sql_query = "INSERT INTO products_manual (upc, name, image, price, employee, open_date, update_date, last_sold) VALUES (" + \
                '"' + upc + '", ' + \
                '"' + product_name + '", ' + \
                '"' + imgpath + '", ' + \
                '"' + str(product_price) + '", ' + \
                '"' + employee + '", ' + \
                '"' + today.strftime("%Y-%m-%d") + '", ' + \
                '"' + today.strftime("%Y-%m-%d") + '", ' + \
                '"' + today.strftime("%Y-%m-%d") + '"' + \
                ");"
            print(sql_query)
            cur.execute(sql_query)
            conn.commit()
            
            
            try:
                addFBListing = request.form["addFBListing"]
            except:
                addFBListing = 'off'
                
            if addFBListing == 'on':
                link_to_mp_post = uploadToMP(product_name, product_price, "", upc, "", "New", imgpath)
                

            add_to_square(product_name, upc, str(product_price) + "00", employee, 0)
            print('Posted to Square')

            itemCounter = update_details(upc, product_name, product_price, 0, "0", employee, link_to_mp_post['link'], link_to_mp_post['imglink'], "")
            print('Posted to Catalog')

            count = itemCounter["count"]
            zpl = printlabel(upc, product_name, product_price, "0")
            
            if "." in request.form["Price"]:
                add_to_square(product_name, upc, request.form["Price"], employee, 0)
            else:
                add_to_square(product_name, upc, request.form["Price"] + "00", employee, 0)
            
            imgname = open_acrobat_print(printlabels)
            
            return render_template('add.html', count=count, zpl=zpl, label=imgname, upc=upcDetails["upc"], vender=vender)
    else:
        vender = request.args.get('vender')
        # upc = request.args.get('upc')
        return render_template('add.html', count=0, zpl="", label="", upc=upcDetails["upc"], vender=vender)
    
@app.route('/updateDiscount', methods = ['GET', 'POST'])
def updateDiscount():
    if(request.method=='POST'):
        
        value = request.form['discounts']
        pk = request.form['pk']
        
        categories = json.load(open(json_url))
        categories[pk]=value
        
        print(categories)
        # Serializing json 
        json_object = json.dumps(categories, indent = 4)
        
        # Writing to sample.json
        with open(json_url, "w") as outfile:
            outfile.write(json_object)

        return render_template('configuration.html', categories=categories)     

@app.route('/configuration', methods = ['GET', 'POST'])            
def configuration():
    
    categories = json.load(open(json_url))
    config = json.load(open(config_url))
    
    if(request.method=='POST'):
        
        config['Images Count'] = request.form['ImagesCount']
        config['Cover Image Path'] = request.form['CoverImagePath']
        config['Schedule'] = request.form['Schedule']
        config['Delay'] = request.form['Delay']
        config['CSEID'] = request.form['CSEID']
        config['GoogleAPI'] = request.form['GoogleAPI']
        config['InstaPrice'] = request.form['InstaPrice']
        
        print(config)
        # Serializing json 
        json_object = json.dumps(config, indent = 4)
        
        # Writing to sample.json
        with open(config_url, "w") as outfile:
            outfile.write(json_object)

        return render_template('configuration.html', categories=categories, config=config)     

    return render_template('configuration.html', categories=categories, config=config)     
 
@app.route('/old', methods = ['GET', 'POST'])
def homepage():
    
    global upcDetails
    
    if request.method == "POST":
        upc=request.form["upc"]
        try:
            stock=request.form["Stock"]
        except:
            stock=0
        try:
            disc=request.form["Discount"]  
        except:
            disc=0
        
        try:
            printlabels = int(request.form["printlabels"])
        except:
            printlabels = 1
             
        employee=request.form["VendorName"]   
        amazon_price =0 
        walmart_price = 0
        target_product_price=0
        
        

        
        if request.form["btn"] == 'Fetch Details':  
            # starting time
            start = time.time()
            
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            tasks = get_price_name(upc, stock, disc, employee), get_walamart_price(upc)
            upcDetails, upcDetailsWalmart = loop.run_until_complete(asyncio.gather(*tasks))
            # tasks = get_price_name(upc, stock, disc, employee), get_walamart_price(upc), get_amazon_price(upc)
            # upcDetails, upcDetailsWalmart, upcDetailsAmazon = loop.run_until_complete(asyncio.gather(*tasks))
            print("get_price_name()={upcDetails}, get_walamart_price()={upcDetailsWalmart}".format(**vars()))
            loop.close()
            
            # upcDetails=get_price_name(upc, stock, disc, employee)   
            lowest_price = 0
            upcDetails = upcDetails
            try:                
                target_product_price = upcDetails["product_price"]
                lowest_price = float(target_product_price)
                upcDetails["lowest_price"] = lowest_price
            except:
                target_product_price = 0
                
            try:                
                walmart_price = upcDetailsWalmart["product_price"]
                if float(target_product_price) == 0:
                    lowest_price = float(walmart_price)
                    upcDetails["lowest_price"] = lowest_price
                elif float(walmart_price) > 0:
                    if float(float(target_product_price) - float(walmart_price)) > 0:
                        lowest_price = float(walmart_price)
                        upcDetails["lowest_price"] = lowest_price
            except:
                walmart_price = 0
            
            # try:                
            #     amazon_price = upcDetailsAmazon["price"].replace('$','')
            #     if float(lowest_price) == 0:
            #         lowest_price = float(amazon_price)
            #     elif float(amazon_price) > 0:
            #         if float(float(lowest_price) - float(amazon_price)) > 0:
            #             lowest_price = float(amazon_price)                
            # except:
            #     amazon_price = 0    
            
            if 'Not Found' in upcDetails["product_name"]:
                print('Not found in Target')
                # upcDetails = get_walamart_price(upc)
                upcDetails = upcDetailsWalmart
                walmart_price = upcDetails["product_price"]
                
                target_product_price="0"
                                

                # if 'Not Found' in upcDetails["product_name"]:
                #     print('Not found in Walmart')
                #     # data = get_amazon_price(upc)
                #     data = upcDetailsAmazon
                #     try:
                #         amazon_price = data['price']
                #         product_name = data['name']
                #     except:
                #         print('Not found in Amazon')
                #         amazon_price = 0
                #         product_name = 'NOT FOUND'
                        
                #     upcDetails = {
                #                 "upc":upc,
                #                 "product_name":product_name,
                #                 "product_price":lowest_price,
                #                 "product_image" : 'Test_default.png',
                #                 "product_description" : product_name,
                #                 "product_category": 'Select'
                #                 }
            #     else:
            #         amazon_price =0
            # else:
            #     walmart_price =0

             # end time
            end = time.time()
            
            runtime = end - start
            dur='%.2f' % runtime
            
                           
            for filename in os.listdir('static/'):
                if filename.startswith('label'):  # not to remove other images
                    imgname = filename
            
           
            categories = json.load(open(json_url))    
            categoryFoundFlag = 0
            discountPercent = 0
            for name, discount in categories.items():
                if name in upcDetails["product_category"]:  
                    categoryFoundFlag = 1   
                    discountPercent = discount
                       
            return render_template('home.html', 
                                   lowest_price = lowest_price,
                                   details=upcDetails, 
                                   duration=dur, 
                                   counter=0,  
                                   zpl="", 
                                   label=imgname, 
                                   target=target_product_price, 
                                   walmart=walmart_price, 
                                   amazon=amazon_price,
                                   categories=categories,
                                   categoryFoundFlag=categoryFoundFlag,
                                   discount = discountPercent)
            
        elif request.form["btn"] == 'Print Label':
            # starting time
            
            start = time.time()
            
            product_name = upcDetails["product_name"]
            product_price = upcDetails["lowest_price"]
            product_description = upcDetails["product_description"]
            product_category = upcDetails["product_category"]
            

            try:
                addFBListing = request.form["addFBListing"]
            except:
                addFBListing = 'off'


            try:
                imgpath = upcDetails["product_image"]
            except:
                imgpath="Test1.png"
                
            link_img_link={}
            link_img_link['link'] = 'https://origamiitlab.com/wp-content/uploads/2018/12/bird.png'
            link_img_link['imglink'] = 'https://origamiitlab.com/wp-content/uploads/2018/12/bird.png'

            try: 
                price1 = float(product_price.replace("$", ""))
            except:
                price1 = float(product_price)

            if disc=="" or disc=='0':
                discountedPrice=product_price
            else:    
                disc1=float(disc)
                discP = price1 - ((price1*disc1)/100)
                discountedPrice = '%.2f' % discP

            product_price1 =  round(float(discountedPrice))  
            
            if(addFBListing=='on'):
                              
                link_img_link = uploadToMP(product_name, product_price1, product_description, upc, "Miscellaneous", "New", imgpath)
                

            itemCounter = update_details(upc, 
                                         product_name, product_price, stock, disc, employee, link_img_link['link'], link_img_link['imglink'], product_description)
            
            count = itemCounter["count"]
            zpl = printlabel(upc, product_name, product_price, disc)

            add_to_square(product_name, upc, str(product_price) + "00", employee, stock)

            
            try:
                printlabels = int(request.form["printlabels"])
            except:
                printlabels = 1
            
            imgname = open_acrobat_print(printlabels)
            
            
            
            # end time
            end = time.time()
            
            runtime = end - start
            print('runtime')
            print(runtime)
            categories = json.load(open(json_url))
            categoryFoundFlag = 0
            discountPercent = 0
            for name, discount in categories.items():
                if name in upcDetails["product_category"]:  
                    categoryFoundFlag = 1
                    discountPercent = discount
                    
            default_details =   {"upc":upc,
                                    "product_name":product_name,
                                    "product_price":product_price,
                                    "product_category": product_category,
                                }
            return render_template('home.html', 
                                   details=default_details, 
                                   duration=runtime, 
                                   counter=count, 
                                   zpl=zpl, 
                                   label=imgname, 
                                   walmart=0, 
                                   amazon=0, 
                                   categories=categories,
                                   categoryFoundFlag=categoryFoundFlag,
                                   discount = discountPercent)
        
        
        else:
            print('Report bad parameter')
    else: 
        categories = json.load(open(json_url))   
        categoryFoundFlag = 0
        discountPercent = 0
        
        default_details = {"upc":"",
                       "product_name":"Title",
                       "product_price":"$0",
                       "product_category": "Select"
                    }
        
        for name, discount in categories.items():
            if name in default_details["product_category"]:  
                categoryFoundFlag = 1
                discountPercent = discount
                
        
        return render_template('home.html', 
                               details=default_details,
                               duration=0,counter=0,
                               walmart=0,
                               amazon=0,
                               categories=categories,
                               categoryFoundFlag=categoryFoundFlag,
                               discount=discountPercent)

@app.route('/add_to_marketplace', methods = ['GET', 'POST'])
def add_to_marketplace():
    
    config = json.load(open(config_url))
    delay = int(config['Delay'])
    delayInMin = int(delay * 60)

    if request.method == "POST":

        if request.form["btn"] == 'Publish':
            product_name = request.form["Name"]  
            product_price = '$'+request.form["Price"]  
            description = request.form["Description"] 
            category = "Miscellaneous"
            condition = "New"         
            upc = datetime.datetime.now().strftime('%Y%m%d%H%M')
            CoverImagePath = request.form["CoverImagePath"]
            ImageCount = config["Images Count"]
            
            path, dirs, files = next(os.walk(CoverImagePath))
            file_count = len(files)
            
            for i in range(int(int(file_count)/int(ImageCount))):
                publish(product_name, product_price, description, upc, category, condition, CoverImagePath, ImageCount)
                time.sleep(delayInMin)     
            
            return render_template('addtomarketplace.html', config=config)
 
    return render_template('addtomarketplace.html', config=config)  
    
@app.route('/amazon', methods = ['GET', 'POST'])
def amazon():
    
    global upcDetailsAmazon

    config = json.load(open(config_url))
    InstaPrice = int(config['InstaPrice'])
    
    if request.method == "POST":
        
        try:
            stock=request.form["Stock"]
        except:
            stock=0
        try:
            disc=request.form["Discount"]  
        except:
            disc=0
        
        try:
            printlabels = int(request.form["printlabels"])
        except:
            printlabels = 1
             
        employee=request.form["VendorName"]   
        amazon_price = 0 
        amazon_link = request.form["AmazonLink"]
        
        data = get_amazon_price_by_link(amazon_link)
        
        try:            
            product_image_url = str(data['image'])        
            urllib.request.urlretrieve(product_image_url, "C:\\Users\\justs\\OneDrive\\Pictures\\ProductImages\\Test1.png")
            product_image_name= 'Test1.png'
        except:
            product_image_name= 'Test_default.png'


        if data['price']['withdeal'] is not None:
            amazon_price = data['price']['withdeal']                  
        elif data['deal'] is not None:
            amazon_price = data['deal']  
        elif data['sale'] is not None:
            amazon_price = data['sale']                    
        else:
            print('Not found in Amazon')
            amazon_price = 0

        # try:
        #     amazon_price = data['deal']
        # except:
        #     print('Not found in Amazon')
        #     amazon_price = data['price']

        try:
            # amazon_price = data['price']
            lowest_price = amazon_price
            product_name = data['name']
        except:
            print('Not found in Amazon')
            amazon_price = 0
            lowest_price = amazon_price
            product_name = 'NOT FOUND'
            
        upcDetailsAmazon = {
                    "product_name":product_name,
                    "product_price":amazon_price,
                    "product_image" : product_image_name,
                    "product_description" : product_name,
                    "product_category": 'Select'
                    }
        
        if request.form["btn"] == 'Fetch Details':  
            # starting time
            start = time.time()
            
             
            for filename in os.listdir('static/'):
                if filename.startswith('label'):  # not to remove other images
                    imgname = filename
            
           
            categories = json.load(open(json_url))    
            categoryFoundFlag = 0
            discountPercent = 0


            end = time.time()
            
            runtime = end - start
            dur = '%.2f' % runtime
                       
            return render_template('amazon.html', 
                                   lowest_price = lowest_price,
                                   details=upcDetailsAmazon, 
                                   duration=dur, 
                                   counter=0,  
                                   zpl='', 
                                   label=imgname, 
                                   amazon=amazon_price,
                                   categories=categories,
                                   categoryFoundFlag=categoryFoundFlag,
                                   discount = discountPercent)
            
            
            
            
        elif request.form["btn"] == 'Print Label':
            # starting time
            
            start = time.time()
                        
            upc = datetime.datetime.now().strftime('%Y%m%d%H%M')
            product_name = upcDetailsAmazon["product_name"]
            product_price = upcDetailsAmazon["product_price"]
            product_description = upcDetailsAmazon["product_description"]
            product_category = 'Select'
            

            try:
                addFBListing = request.form["addFBListing"]
            except:
                addFBListing = 'off'


            try:
                imgpath = upcDetailsAmazon["product_image"]
            except:
                imgpath="Test1.png"
                
            link_img_link={}
            link_img_link['link'] = 'https://origamiitlab.com/wp-content/uploads/2018/12/bird.png'
            link_img_link['imglink'] = 'https://origamiitlab.com/wp-content/uploads/2018/12/bird.png'
                
            price1 = float(product_price.replace("$", ""))
        
            if disc=="" or disc=='0':
                discountedPrice=float(product_price.replace("$", ""))
            else:    
                disc1=float(disc)
                discP = price1 - ((price1*disc1)/100)
                discountedPrice = '%.2f' % discP

            product_price1 =  round(float(discountedPrice))  

            if(addFBListing=='on'):                              
                link_img_link = uploadToMP(product_name, product_price1, product_description, upc, "Miscellaneous", "New", imgpath)
                
                if int(product_price1)>int(InstaPrice):
                    writeInstaJson(upc, product_name, product_description, imgpath, product_price, disc, stock, employee, link_img_link['link'])

            itemCounter = update_details(upc, 
                                         product_name, 
                                         product_price, stock, disc, employee, link_img_link['link'], link_img_link['imglink'], product_description)
            
            count = itemCounter["count"]
            zpl = printlabel(upc, product_name, product_price, disc)

            add_to_square(product_name, upc, product_price + "00", employee, stock)

            
            try:
                printlabels = int(request.form["printlabels"])
            except:
                printlabels = 1
            
            imgname = open_acrobat_print(printlabels)
            
            
            
            # end time
            end = time.time()
            
            runtime = end - start
            print('runtime')
            print(runtime)
            categories = json.load(open(json_url))
            categoryFoundFlag = 0
            discountPercent = 0
            for name, discount in categories.items():
                if name in upcDetailsAmazon["product_category"]:  
                    categoryFoundFlag = 1
                    discountPercent = discount
                    
            default_details =   {
                                    "product_name":product_name,
                                    "product_price":product_price,
                                    "product_category": product_category,
                                }
            return render_template('amazon.html', 
                                   details=default_details, 
                                   duration=runtime, 
                                   counter=count, 
                                   zpl=zpl, 
                                   label=imgname, 
                                   walmart=0, 
                                   amazon=0, 
                                   categories=categories,
                                   categoryFoundFlag=categoryFoundFlag,
                                   discount = discountPercent)
        
        
        else:
            print('Report bad parameter')
    else: 
        categories = json.load(open(json_url))   
        categoryFoundFlag = 0
        discountPercent = 0
        
        default_details = {
                       "product_name":"Title",
                       "product_price":"$0",
                       "product_category": "Select"
                    }
        
        for name, discount in categories.items():
            if name in default_details["product_category"]:  
                categoryFoundFlag = 1
                discountPercent = discount
                
        
        return render_template('amazon.html', 
                               details=default_details,
                               duration=0,counter=0,
                               walmart=0,
                               amazon=0,
                               categories=categories,
                               categoryFoundFlag=categoryFoundFlag,
                               discount=discountPercent)

@app.route('/amazon_add_multiple', methods = ['GET', 'POST'])
def amazon_add_multiple():
    
    config = json.load(open(config_url))
    delay = int(config['Delay'])
    delayInMin = int(delay * 60)

    if request.method == "POST":

        if request.form["btn"] == 'Publish':

            description = request.form["Description"] 
            category = "Miscellaneous"
            condition = "New"         
            upc = datetime.datetime.now().strftime('%Y%m%d%H%M')
            SheetPath = request.form["SheetPath"]
            
            
            wb = load_workbook(SheetPath)
            sheet = wb.worksheets[0]

            row_count = sheet.max_row
            
            for i in range(2, row_count):
                
                product_name = sheet.cell(row = i, column = 2).value
                product_price = '$'+ str(sheet.cell(row = i, column = 1).value)
                amazon_link = sheet.cell(row = i, column = 3).value
                
                data = get_amazon_price_by_link(amazon_link)
        
                try:            
                    product_image_url = str(data['image'])        
                    urllib.request.urlretrieve(product_image_url, "C:\\Users\\justs\\OneDrive\\Pictures\\ProductImages\\Test1.png")
                    product_image_name= 'Test1.png'
                except:
                    product_image_name= 'Test_default.png'

                
                # if data['price'] is not None:
                #     product_price = data['price']                    
                # elif data['deal'] is not None:
                #     product_price = data['deal']  
                # elif data['sale'] is not None:
                #     product_price = data['sale']                    
                # else:
                #     print('Not found in Amazon')
                #     product_price = 0

                # product_name = data['name']
                uploadToMP(product_name, product_price, description, upc, category, condition, product_image_name)
                time.sleep(delayInMin)     
            
            return render_template('addtomarketfromamazon.html', config=config)
 
    return render_template('addtomarketfromamazon.html', config=config)  

def writeInstaJson(upc, product_name, product_description, product_image, product_price, disc, stock, employee, link_to_mp_post):

    
    # python object to be appended
    new_data = {
        'upc': upc,
        'product_name': product_name,
        'product_description': product_description,
        'product_image': product_image,
        'product_price': product_price, 
        'disc': disc,
        'stock': stock, 
        'employee': employee,
        'link_to_mp_post': link_to_mp_post
        }

    with open(data_ig_url,'r+') as file:
        # First we load existing data into a dict.
        file_data = json.load(file)
        # check if UPC exist
        for upc_no in file_data["upc_details"]:
            if str(upc) == str(upc_no['upc']):
                return
            
        # Join new_data with file_data inside upc_details
        file_data["upc_details"].append(new_data)
        # Sets file's current position at offset.
        file.seek(0)
        # convert back to json.
        json.dump(file_data, file, indent = 4)

    
    return

def writeAllDetailsInCSV(upc, product_name, product_description, product_image, product_price, disc, stock, employee):

    
    # python object to be appended
    new_data = {
        'upc': upc,
        'product_name': product_name,
        'product_description': product_description,
        'product_image': product_image,
        'product_price': product_price, 
        'disc': disc,
        'stock': stock, 
        'employee': employee
        }

    if os.path.isfile(data_date_url):
        # checks if file exists
        print ("File exists and is readable")
    else:
        print ("Either file is missing or is not readable, creating file...")
        with open(data_date_url, 'w') as db_file:
            db_file.write(json.dumps({"upc_details":[]}))


    with open(data_date_url,'r+') as file:
        # First we load existing data into a dict.
        file_data = json.load(file)
        # check if UPC exist
        for upc_no in file_data["upc_details"]:
            if upc == upc_no['upc']:
                return
            
        # Join new_data with file_data inside upc_details
        file_data["upc_details"].append(new_data)
        # Sets file's current position at offset.
        file.seek(0)
        # convert back to json.
        json.dump(file_data, file, indent = 4)

    
    return

def delete_upc(upc):
    print("deleteing upc" + str(upc))
    data_array = json.load(open(data_url))
    df_data = pd.DataFrame(data_array['upc_details'])
    df_data.drop(df_data[df_data['upc'] == upc].index, inplace=True)
    data_dict = df_data.to_dict('records')
    # convert df to dict
    json_f = dict()
    json_f['upc_details'] = data_dict
    # df to json and save it to json file
    result = json.dumps(json_f, indent=4)
    jsonFile = open(data_url, "w")
    jsonFile.write(result)
    jsonFile.close()

def upload_products():

    config = json.load(open(config_url))
    data_array = json.load(open(data_url))
    delay = int(config['Delay'])
    delayInMin = int(delay * 60)
    InstaPrice = int(config['InstaPrice'])


    # counter = 0
    # sql_query = "SELECT * FROM " + table_name

    for row in data_array['upc_details']:        

        upc = row['upc']
        product_name = row['product_name']
        product_description = row['product_description']
        product_image = row['product_image']
        product_price = row['product_price']
        disc = row['disc']
        stock = row['stock']
        employee = row['employee']

    
        try:                  
            urllib.request.urlretrieve(product_image, "C:\\Users\\justs\\OneDrive\\Pictures\\ProductImages\\Test1.png")   
            product_image_name= 'Test1.png'
        except:
            product_image = 'https://origamiitlab.com/wp-content/uploads/2018/12/bird.png'
            product_image_name= 'Test_default.png'
        

        try: 
            price1 = float(product_price.replace("$", ""))
        except:
            price1 = float(product_price)

        if disc=="" or disc=='0':
            discountedPrice=product_price
        else:    
            disc1=float(disc)
            discP = price1 - ((price1*disc1)/100)
            discountedPrice = '%.2f' % discP

        product_price1 =  round(float(discountedPrice)) 

        try:    
            print(product_name + " " + str(upc))    
            link_to_mp_post = uploadToMP(product_name, product_price1, product_description, upc, "Miscellaneous", "New", product_image_name)
            print('Posted to MP ' + link_to_mp_post['link'])
            if link_to_mp_post['link']=='':
                delete_upc(upc)


            if int(product_price1)>int(InstaPrice):
                writeInstaJson(upc, product_name, product_description, product_image, product_price, disc, stock, employee, link_to_mp_post)

            itemCounter = update_details(upc, product_name, product_price, stock, disc, employee, link_to_mp_post['link'], product_image, product_description)
            print('Posted to Catalog')

            count = itemCounter["count"]

            add_to_square(product_name, upc, str(product_price) + "00", employee, stock)
            print('Posted to Square')


        except Exception as e: 
            print(str(e))
            # os.system("taskkill /im chromedriver.exe /F")
            # os.system("taskkill /im chrome.exe /F")
            print('Failed UPC - ' + upc)


    return

def upload_products_to_ig():    
    
    data_array = json.load(open(data_ig_url))
    counter = 0

    for row in data_array['upc_details']:

        if counter==24:
            break

        upc = row['upc']
        product_name = row['product_name']
        product_description = row['product_description']
        product_image = row['product_image']
        product_price = row['product_price']
        disc = row['disc']
        stock = row['stock']
        employee = row['employee']
        link_to_mp_post = row['link_to_mp_post']


        try: 
            price1 = float(product_price.replace("$", ""))
        except:
            price1 = float(product_price)

        if disc=="" or disc=='0':
            discountedPrice=product_price
        else:    
            disc1=float(disc)
            discP = price1 - ((price1*disc1)/100)
            discountedPrice = '%.2f' % discP

        product_price1 =  round(float(discountedPrice)) 
        description = str(product_name) + """
        
        """ + str(link_to_mp_post) + """
        
        """ +'PRICE = $' + str(product_price1) 
        

        if(len(product_name)>109):
            description=product_name[0:2200]
        else:
            pass

        postInstagram(product_image, description)

        print('Posted to IG')

        counter = counter + 1

    return

table_name_products = ""

@app.route('/', methods = ['GET', 'POST'])
def target():
    conn = sqlite3.connect('mydb.db')
    cur = conn.cursor()
    global table_name, table_name_products
    table_name = "data" + \
                    (datetime.datetime.today().strftime("%Y%m%d"))
    sql_query = "CREATE TABLE IF NOT EXISTS " + table_name + "_scaned" + \
                    " ( id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, " + \
                        "upc TEXT," + \
                        "name TEXT," + \
                        "description TEXT," + \
                        "image TEXT," + \
                        "price REAL," + \
                        "category TEXT," + \
                        "disc REAL," + \
                        "stock INTEGER," + \
                        "employee TEXT);"
    print(sql_query)
    cur.execute(sql_query)
    conn.commit()    
    sql_query = "CREATE TABLE IF NOT EXISTS " + table_name + "_printed" + \
                    " ( id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, " + \
                        "upc TEXT," + \
                        "name TEXT," + \
                        "description TEXT," + \
                        "image TEXT," + \
                        "price REAL," + \
                        "category TEXT," + \
                        "disc REAL," + \
                        "stock INTEGER," + \
                        "employee TEXT);"
    print(sql_query)
    cur.execute(sql_query)
    conn.commit()    
    global upcDetails
    
    if request.method == "POST":
        link=str(request.form["TargetLink"])
        try:
            stock=request.form["Stock"]
        except:
            stock=0
        try:
            disc=request.form["Discount"]  
        except:
            disc=0
        

        if '' in request.form["printlabels"]:
            printlabels = 1
        else:
            printlabels = int(request.form["printlabels"])

             
        employee=request.form["VendorName"]  
        amazon_price =0 
        walmart_price = 0
        target_product_price=0
        
        categories = {}
        sql_query = "SELECT category_name, discount FROM discounts"
        print(sql_query)
        results = cur.execute(sql_query).fetchall()
        print(results)
        for row in results:
            categories[row[0]] = row[1]

        
        if request.form["btn"] == 'Fetch Details':  
            
            start = time.time()
            
            product_url = 'Not Found'
            product_upc = link
            product_name = 'Not Found'
            product_description = 'Not Found'
            product_image = 'Not Found'
            product_category = 'Not Found'
            product_price = 'Not Found'
            product_disc = 'Not Found'
            product_stock = 'Not Found'
            product_employee = 'Not Found'
            product_open = 'Not Found'
            product_update = 'Not Found'
            # product_close = 'Not Found'
            # product_is_available = 'Not Found'
            product_last_sold = 'Not Found'
            product_last_price = 'Not Found'
            upcDetails = {
                "url": product_url,
                "upc": link,
                "name": product_name,
                "description": product_description,
                "image": product_image,
                "category": product_category,
                "price": product_price,
                "disc": product_disc,
                "stock": product_stock,
                "employee": product_employee,
                "open": product_open,
                "update": product_update,
                "last_sold": product_last_sold,
                "last_price": product_last_price,
                # "upc": product_upc,
                # "is_available": product_is_available
            }
            
            lowest_price = 0
            categoryFoundFlag = 0
            discountPercent = 0
            
            
            
            tcin_category = get_tcin_upc(link)
            if (isinstance(tcin_category, int) and tcin_category > 300) or tcin_category['tcin'] == 'Not Found' or tcin_category['category'] == 'Not Found':
                tcin = category = "Not Found"
                table_name_products = category + "_products"
                table_name_products = table_name_products.replace(" ", "")
                sql_query = "SELECT * FROM products_manual WHERE upc=" + '"' + link + '"'
                results = cur.execute(sql_query).fetchall()
                if len(results):
                    if results[0][15] == "0":
                        flash("Not found product's tcin with current upc on target.com!")
                        flash("This product is not available yet!")
                    else:
                        flash("Not found product's tcin with current upc on target.com!")
                        flash("This product is not available now!")
                        id = results[0][0]
                        sql_query = "UPDATE products_manual SET " + \
                            "update_date=" + '"' + today.strftime('%Y-%m-%d') + '", ' + \
                            "is_available=0" + \
                            " WHERE id=" + str(id)
                        cur.execute(sql_query)
                        conn.commit()
                    product_url = results[0][1]
                    # product_upc = results[0][3]
                    product_name = results[0][4]
                    product_description = results[0][5]
                    product_image = results[0][6]
                    product_category = results[0][7]
                    product_price = results[0][8]
                    product_disc = results[0][9]
                    product_stock = results[0][10]
                    product_employee = results[0][11]
                    product_open = results[0][12]
                    product_update = results[0][13]
                    # product_close = results[0][14]
                    # product_is_available = results[0][15]
                    product_last_sold = results[0][16]
                    product_last_price = results[0][17]
                    upcDetails = {
                        "url": product_url,
                        "tcin": tcin,
                        "upc": link,
                        "name": product_name,
                        "description": product_description,
                        "image": product_image,
                        "category": product_category,
                        "price": product_price,
                        "discount": product_disc,
                        "stock": product_stock,
                        "employee": product_employee,
                        "open": product_open,
                        "update": product_update,
                        "last_sold": product_last_sold,
                        "last_price": product_last_price,
                        # "upc": product_upc,
                        # "is_available": product_is_available
                    }
                    
                else:
                    flash("Not found product's tcin with current upc on target.com!")
                    flash("It may be new product!")
                    return redirect(url_for('add_produtcs', vender=employee, upc=link))
                    
            else:
                tcin = tcin_category['tcin']
                category = tcin_category['category']
                table_name_products = category + "_products"
                table_name_products = table_name_products.replace(" ", "")
                sql_query = "SELECT discount FROM discounts WHERE category_name=" + '"' + category + '"'
                print(sql_query)
                results = cur.execute(sql_query).fetchall()
                print(results)
                if len(results):
                    categoryFoundFlag = 1
                    discountPercent = results[0][0]
                sql_query = "SELECT * FROM " + table_name_products + \
                    " WHERE tcin=" + '"' + str(tcin) + '"'
                results = cur.execute(sql_query).fetchall()
                if len(results):
                    product_url = results[0][1]
                    # product_upc = results[0][3]
                    product_name = results[0][4]
                    product_description = results[0][5]
                    product_image = results[0][6]
                    product_category = results[0][7]
                    product_price = results[0][8]
                    product_disc = results[0][9]
                    product_stock = results[0][10]
                    product_employee = results[0][11]
                    product_open = results[0][12]
                    product_update = results[0][13]
                    # product_close = results[0][14]
                    # product_is_available = results[0][15]
                    product_last_sold = results[0][16]
                    product_last_price = results[0][17]
                    
                    upcDetails = {
                        "url": product_url,
                        "tcin": tcin,
                        "upc": link,
                        "name": product_name,
                        "description": product_description,
                        "image": product_image,
                        "category": product_category,
                        "price": product_price,
                        "discount": product_disc,
                        "stock": product_stock,
                        "employee": product_employee,
                        "open": product_open,
                        "update": product_update,
                        "last_sold": product_last_sold,
                        "last_price": product_last_price,
                        # "upc": product_upc,
                        # "is_available": product_is_available
                    }
                    
                    upcDetails["employee"] = employee
                    
                    try:                
                        target_product_price = upcDetails["price"]
                        lowest_price = float(target_product_price)
                    except:
                        target_product_price = 0
                        lowest_price = 0
                else:
                    flash("Not found product's tcin with current upc on target.com!")
                    flash("It is new product!")
                    product_info = get_products_tcin(tcin)
                    upcDetails = {
                        "url": product_info['url'],
                        "tcin": tcin,
                        "upc": link,
                        "name": product_info['name'],
                        "description": product_info['description'],
                        "image": product_info['image'],
                        "category": category,
                        "price": product_info['price_min'],
                        "discount": str(discountPercent),
                        "employee": employee
                    }
                    sql_query = 'INSERT INTO ' + table_name_products + ' (url, tcin, name, description, image, category, price, disc, employee, open_date, update_date) ' + \
                        " VALUES (" + \
                        '"' + upcDetails['url'] + '", ' + \
                        '"' + upcDetails['tcin'] + '", ' + \
                        '"' + upcDetails['name'].replace("\"", " ") + '", ' + \
                        '"' + upcDetails['description'].replace("\"", " ") + '", ' + \
                        '"' + upcDetails['image'] + '", ' + \
                        '"' + upcDetails['category'] + '", ' + \
                        '"' + upcDetails['price'] + '", ' + \
                        '"' + upcDetails['discount'] + '", ' + \
                        '"' + employee + '", ' + \
                        '"' + today.strftime("%Y-%m-%d") + '", ' + \
                        '"' + today.strftime("%Y-%m-%d") + '"' + ");"
                    print(sql_query)
                    cur.execute(sql_query)
                    conn.commit()
                print(upcDetails)
            sql_query = "INSERT INTO " + table_name + "_scaned" + " (upc, name, description, image, price, category, disc, stock, employee) VALUES (" + \
                            '"' + upcDetails["upc"] + '", ' + \
                            '"' + upcDetails["name"] + '", ' + \
                            '"' + upcDetails['description'] + '", ' + \
                            '"' + upcDetails["image"] + '", ' + \
                            '"' + upcDetails["price"] + '", ' + \
                            '"' + upcDetails["category"] + '", ' + \
                            '"' + upcDetails["discount"] + '", ' + \
                            '"' + str(stock) + '", ' + \
                            '"' + employee + '");'
            print(sql_query)
            cur.execute(sql_query)
            conn.commit()
                    
                    
                    
                
            for filename in os.listdir('static/'):
                if filename.startswith('label'):  # not to remove other images
                    imgname = filename
            
           
            # categories = json.load(open(json_url))  
            # print(categories)  
            # categoryFoundFlag = 0
            # discountPercent = 0
            # for name, discount in categories.items():
            #     if name in upcDetails["product_category"]:  
            #         categoryFoundFlag = 1   
            #         discountPercent = discount
            


            # auto start
            # upc = upcDetails["upc"]
            # product_name = upcDetails["product_name"]
            # product_price = upcDetails["lowest_price"]
            # product_description = upcDetails["product_description"]
            # product_category = upcDetails["product_category"]
            # product_image = upcDetails["product_image"]
            # disc = upcDetails["discount"]
            # try:
            #    stock = product_stock
            # except:
            #     pass
            # try:
            #     employee = product_employee
            # except:
            #     pass
            
            # if not 'Not Found' in upcDetails["product_name"]:
            #     sql_query = "INSERT INTO " + table_name + "_scaned" + " (upc, name, description, image, price, category, disc, stock, employee) VALUES (" + \
            #                 '"' + upcDetails["upc"] + '", ' + \
            #                 '"' + upcDetails["product_name"] + '", ' + \
            #                 '"' + upcDetails['product_description'] + '", ' + \
            #                 '"' + upcDetails["product_image"] + '", ' + \
            #                 str(upcDetails["lowest_price"]) + ", " + \
            #                 '"' + category + '", ' + \
            #                 str(disc) + ", " + \
            #                 str(stock) + ", " + \
            #                 '"' + employee + '");'
            #     print(sql_query)
            #     cur.execute(sql_query)
            #     conn.commit()
            # elif not product_is_available:  # not available product
            #     pass
            # else:   # new product
            #     flash("Please add this product into your database!")
            #     if tcin == 'Not Found':
            #         return redirect(url_for('add_produtcs', vender=employee, upc=upcDetails["upc"]))
            #     else:
            #         upcDetails = get_products_tcin(tcin)
            #         sql_query = "INSERT INTO " + table_name + "_scaned" + " (upc, name, description, image, price, category, disc, stock, employee) VALUES (" + \
            #                 '"' + upcDetails["upc"] + '", ' + \
            #                 '"' + upcDetails["name"] + '", ' + \
            #                 '"' + upcDetails['description'] + '", ' + \
            #                 '"' + upcDetails["image"] + '", ' + \
            #                 '"' + str(upcDetails["price_min"]) + '", ' + \
            #                 '"' + category + '", ' + \
            #                 str(disc) + ", " + \
            #                 str(stock) + ", " + \
            #                 '"' + employee + '");'
            #         print(sql_query)
            #         cur.execute(sql_query)
            #         conn.commit()
            #         sql_query = "INSERT INTO products (url, tcin, upc, name, description, image, category, price, open_date, update_date, last_sold, employee) VALUES (" + \
            #             '"' + upcDetails['url'] + '", ' + \
            #             '"' + tcin + '", ' + \
            #             '"' + upcDetails["upc"] + '", ' + \
            #             '"' + upcDetails["name"] + '", ' + \
            #             '"' + upcDetails['description'] + '", ' + \
            #             '"' + upcDetails["image"] + '", ' + \
            #             '"' + product_category + '", ' + \
            #             '"' + str(upcDetails["price_min"]) + '", ' + \
            #             '"' + str(today.strftime('%Y-%m-%d')) + '", ' + \
            #             '"' + str(today.strftime('%Y-%m-%d')) + '", ' + \
            #             '"' + str(today.strftime('%Y-%m-%d')) + '", ' + \
            #             '"' + employee + '");'
            #         print(sql_query)
            #         cur.execute(sql_query)
            #         conn.commit()

            # writeAllDetailsInCSV(upc, product_name, product_description,
            #                      product_image, product_price, disc, stock, employee)

            # auto end
                
            # end time
            end = time.time()
            runtime = end - start
            dur='%.2f' % runtime
            
            conn.close()
            
            return render_template('target.html', 
                                   lowest_price = lowest_price,
                                   details=upcDetails, 
                                   duration=dur, 
                                   counter=0,  
                                   zpl="", 
                                   label=imgname, 
                                   target=target_product_price,
                                   walmart=walmart_price, 
                                   amazon=amazon_price,
                                   categories=categories,
                                   categoryFoundFlag=categoryFoundFlag,
                                   discount = discountPercent)
            
        elif request.form["btn"] == 'Print Label':
            # starting time
            
            start = time.time()
            
            print(upcDetails)
            
            # if not 'Not Found' in upcDetails["tcin"] and not '' in upcDetails['tcin']:
            sql_query = "INSERT INTO " + table_name + "_printed" + \
                    " (upc, name, description, image, price, category, disc, stock, employee) VALUES (" + \
                    '"' + upcDetails['upc'] + '", ' + \
                    '"' + upcDetails['name'] + '", ' + \
                    '"' + upcDetails['description'] + '", ' + \
                    '"' + upcDetails['image'] + '", ' + \
                    '"' + upcDetails['price'] + '", ' + \
                    '"' + upcDetails['category'] + '", ' + \
                    '"' + str(disc) + '", ' + \
                    '"' + str(stock) + '", ' + \
                    '"' + employee + '");'
            print(sql_query)
            cur.execute(sql_query)
            conn.commit()
            sql_query = "UPDATE " + table_name_products + " SET " + \
                        "stock=" + '"' + str(stock) + '", ' + \
                        "last_sold=" + '"' + datetime.datetime.today().strftime("%Y%m%d%H%M%S") + '", ' + \
                        "last_price=" + '"' + str(upcDetails['price']) + '"' + \
                        " WHERE tcin=" + '"' + upcDetails['tcin'] + '" or upc=' + '"' + link + '"'
            print(sql_query)
            cur.execute(sql_query)
            conn.commit()
            zpl = printlabel(upcDetails['upc'], upcDetails['name'], upcDetails['price'], disc)

            try:
                    printlabels = int(request.form["printlabels"])
            except:
                    printlabels = 1

            imgname = open_acrobat_print(printlabels)

            # else:
                # sql_query = "INSERT INTO " + table_name + "_printed" + \
                #     " (upc, name, description, image, price, category, disc, stock, employee) VALUES (" + \
                #     '"' + upcDetails['upc'] + '", ' + \
                #     '"' + upcDetails['name'] + '", ' + \
                #     '"' + upcDetails['description'] + '", ' + \
                #     '"' + upcDetails['image'] + '", ' + \
                #     '"' + upcDetails['price'] + '", ' + \
                #     '"' + upcDetails['category'] + '", ' + \
                #     '"' + str(disc) + '", ' + \
                #     '"' + str(stock) + '", ' + \
                #     '"' + employee + '");'
                # print(sql_query)
                # cur.execute(sql_query)
                # conn.commit()
                # flash("Wrong product name!")
                # flash("Please try again with another product!")
                # zpl = ""
                # imgname = ""
            # elif not upcDetails['is_available']:  # step 9 - print
            #         sql_query = "INSERT INTO " + table_name + "_printed" + " (upc, name, price, disc, last_sold, last_price) VALUES (" + \
            #                     '"' + upcDetails["upc"] + '", ' + \
            #                     '"' + upcDetails["product_name"] + '", ' + \
            #                     str(upcDetails["product_price"]) + ", " + \
            #                     str(disc) + ", " + \
            #                     '"' + datetime.datetime.today().strftime("%Y-%m-%d") + '", ' + \
            #                     str(upcDetails['lowest_price']) + ");"
            #         print(sql_query)
            #         cur.execute(sql_query)
            #         conn.commit()
            # else:
            #     is_new = 1
            # if is_new:
            #     flash("It is new product")
            #     # return render_template('add.html', count=0, zpl="", label="", upc=upcDetails["upc"])
            #     return redirect(url_for('add_produtcs', vender=employee, upc=upcDetails["upc"]))
            # writeAllDetailsInCSV(upc, product_name, product_description, product_image, product_price, disc, stock, employee)
            
            sql_query = "SELECT * FROM " + table_name + "_printed"
            print(sql_query)
            results = cur.execute(sql_query).fetchall()
            print(results)
            if len(results) % 101 == 100:
                sql_query = "SELECT * FROM " + table_name + "_printed" + " WHERE id BETWEEN " + \
                    str((len(results) / 101) % 101) + \
                    " and " + str(len(results))
                results = cur.execute(sql_query).fetchall()
                for row in results:
                    link_to_mp_post = uploadToMP(
                        row[2], row[5], row[3], row[1], row[6], "New", row[4])
                    add_to_square(row[2], row[1], str(
                        row[5]) + "00", row[9], row[8])
                    print('Posted to Square ' + row[0])
                
            
            categoryFoundFlag = 0
            discountPercent = 0
            sql_query = "SELECT discount FROM discounts WHERE category_name=" + \
                '"' + upcDetails["category"] + '"'
            results = cur.execute(sql_query).fetchall()
            if len(results):
                categoryFoundFlag = 1
                discountPercent = results[0][0]
                    
            # categories = json.load(open(json_url))
            # categoryFoundFlag = 0
            # discountPercent = 0
            # for name, discount in categories.items():
            #     if name in upcDetails["product_category"]:  
            #         categoryFoundFlag = 1
            #         discountPercent = discount
                    
            # default_details =   {   "upc":link,
            #                         "name":product_name,
            #                         "price":product_price,
            #                         "category": product_category,
            #                     }
            # end time
            end = time.time()
            
            runtime = end - start
            print('runtime')
            print(runtime)
            
            conn.close()
            
            return render_template('target.html', 
                                   details=upcDetails, 
                                   duration=runtime, 
                                   counter=0000, 
                                   zpl=zpl, 
                                   label=imgname, 
                                   walmart=0, 
                                   amazon=0, 
                                   categories=categories,
                                   categoryFoundFlag=categoryFoundFlag,
                                   discount = discountPercent)
        
        elif request.form["btn"] == 'Upload Products':

            upload_products()

            categories = json.load(open(json_url))   
            categoryFoundFlag = 0
            discountPercent = 0
            
            default_details = {"upc":"",
                       "product_name":"Title",
                       "product_price":"$0",
                       "product_category": "Select"
                    }

            conn.close()

            return render_template('target.html', 
                               details=default_details,
                               duration=0,counter=0,
                               walmart=0,
                               amazon=0,
                               categories=categories,
                               categoryFoundFlag=categoryFoundFlag,
                               discount=discountPercent)
    
        elif request.form["btn"] == 'Upload 25 to IG':

            upload_products_to_ig()

            categories = json.load(open(json_url))   
            categoryFoundFlag = 0
            discountPercent = 0
            
            default_details = {"upc":"",
                       "product_name":"Title",
                       "product_price":"$0",
                       "product_category": "Select"
                    }

            conn.close()

            return render_template('target.html', 
                               details=default_details,
                               duration=0,counter=0,
                               walmart=0,
                               amazon=0,
                               categories=categories,
                               categoryFoundFlag=categoryFoundFlag,
                               discount=discountPercent)

        else:
            print('Report bad parameter')
            conn.close()
    else:
        try:
            upc = upcDetails["upc"]
        except:
            upc = "" 
        categories = {}
        sql_query = "SELECT category_name, discount FROM discounts"
        results = cur.execute(sql_query).fetchall()
        for row in results:
            categories[row[0]] = row[1]
        # categories = json.load(open(json_url))   
        categoryFoundFlag = 0
        discountPercent = 0
        
        default_details = {"upc": upc,
                       "product_name":"Title",
                       "product_price":"$0",
                       "product_category": "Select"
                    }
        
        sql_query = "SELECT discount FROM discounts WHERE category_name=" + \
            '"' + default_details["product_category"] + '"'
        results = cur.execute(sql_query).fetchall()
        if len(results):
            categoryFoundFlag = 1
            discountPercent = results[0][0]
                
        conn.close()
        return render_template('target.html', 
                               details=default_details,
                               duration=0,
                               counter=0,
                               walmart=0,
                               amazon=0,
                               categories=categories,
                               categoryFoundFlag=categoryFoundFlag,
                               discount=discountPercent)


async def get_upc():
    browser = await pyppeteer.launch(headless=False,
                                    #  executablePath="C:/Program Files/Google/Chrome/Application/chrome.exe",
                                     handleSIGINT=False,
                                     handleSIGTERM=False,
                                     handleSIGHUP=False)
    # endpoint = browser.wsEndpoint()
    # pyppeteer.connect(browserWSEndpoint=endpoint)
    # page = pyppeteer.connect(browserURL='http://127.0.0.1:9222')
    page = await browser.newPage()
    await page.goto("https://www.target.com")
    while True:
        try:
            content = await page.content()
            soup = bs4.BeautifulSoup(content, features="lxml")
            btn = soup.find("button", string="go")
            global upcDetails
            upc = soup.select('input#select')[0]['value']
            print("upc:" + upc)
            if upc:
                upcDetails['upc'] = upc
                target()
                upc = soup.select('input#select')[0]['value'] = ''

        except Exception as err:
            # print(err)
            pass
class myThread (threading.Thread):
    def __init__(self, threadID, name, delayTime):
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name
        self.delayTime = delayTime

    def run(self):
        print(self.name)
        testing()
        print('Done Testing!')

th = myThread(0, 'Start Testing...',  2)
th.start()

if __name__ == '__main__':
    app.run(port=5000, debug=True)
